from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
import traceback

from app.database import get_db
from app.services.stats_service import (
    subject_stats,
    score_distribution,
    grade_overview,
    combination_overview,
    class_comparison,
    student_trend,
    detect_weak_subjects,
    exam_quality_analysis,
    score_segment_stats,
    export_quality_to_excel,
    rank_distribution,
    score_shape_analysis,
    critical_students_stats,
)

router = APIRouter()


@router.get("/subject/{exam_subject_id}")
def get_subject_stats(exam_subject_id: int, db: Session = Depends(get_db)):
    """单科统计"""
    return subject_stats(db, exam_subject_id)


@router.get("/distribution/{exam_subject_id}")
def get_distribution(exam_subject_id: int, step: int = Query(10, ge=5, le=50), db: Session = Depends(get_db)):
    """分数段分布"""
    return score_distribution(db, exam_subject_id, step)


@router.get("/grade/{exam_id}")
def get_grade_overview(exam_id: int, db: Session = Depends(get_db)):
    """年级总览"""
    overview = grade_overview(db, exam_id)
    try:
        combo = combination_overview(db, exam_id)
    except Exception:
        traceback.print_exc()
        combo = {"班级数据": {}, "有赋分": False, "年级汇总": {}}
    overview["班级数据"] = combo.get("班级数据", {})
    overview["有赋分"] = combo.get("有赋分", False)
    overview["年级汇总"] = combo.get("年级汇总", {})
    return overview


@router.get("/class-comparison/{exam_id}")
def get_class_comparison(exam_id: int, db: Session = Depends(get_db)):
    """班级对比"""
    return class_comparison(db, exam_id)


@router.get("/quality-analysis/{exam_id}")
def get_quality_analysis(exam_id: int, db: Session = Depends(get_db)):
    """考试质量分析表"""
    try:
        return exam_quality_analysis(db, exam_id)
    except Exception:
        traceback.print_exc()
        return {"exam_id": exam_id, "classes": [], "subjects": {}}


@router.get("/score-segments/{exam_id}")
def get_score_segments(exam_id: int, db: Session = Depends(get_db)):
    """成绩分段统计"""
    return score_segment_stats(db, exam_id)


@router.get("/export-quality/{exam_id}")
def export_quality_excel(exam_id: int, db: Session = Depends(get_db)):
    """导出考试质量分析和分段统计为 Excel"""
    data = export_quality_to_excel(db, exam_id)
    if not data:
        return {"error": "无数据可导出"}
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=quality_analysis_{exam_id}.xlsx"},
    )


@router.get("/student-trend/{student_id}")
def get_student_trend(student_id: int, db: Session = Depends(get_db)):
    """学生纵向趋势"""
    return student_trend(db, student_id)


@router.get("/weak-subjects/{student_id}")
def get_weak_subjects(student_id: int, db: Session = Depends(get_db)):
    """检测薄弱科目"""
    return detect_weak_subjects(db, student_id)


@router.get("/rank-distribution/{exam_id}")
def get_rank_distribution(
    exam_id: int,
    top_n: int = Query(30, ge=1, le=200),
    bottom_n: int = Query(30, ge=1, le=200),
    db: Session = Depends(get_db)
):
    """排名分布：前x名和后x名中各班级学生的分布"""
    return rank_distribution(db, exam_id, top_n, bottom_n)


@router.get("/score-shape/{exam_id}")
def get_score_shape(exam_id: int, db: Session = Depends(get_db)):
    """分数分布形态分析"""
    return score_shape_analysis(db, exam_id)


@router.get("/critical-students/{exam_id}")
def get_critical_students(exam_id: int, db: Session = Depends(get_db)):
    """临界生统计"""
    return critical_students_stats(db, exam_id)


@router.get("/export-rank-distribution/{exam_id}")
def export_rank_distribution_excel(
    exam_id: int,
    top_n: int = Query(30, ge=1, le=200),
    bottom_n: int = Query(30, ge=1, le=200),
    db: Session = Depends(get_db)
):
    """导出排名分布为Excel"""
    data = rank_distribution(db, exam_id, top_n, bottom_n)
    if not data:
        return {"error": "无数据可导出"}

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "排名分布"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center")

    # 表头
    ws.cell(row=1, column=1, value="班级").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = center

    top_info = data.get("top", {})
    bottom_info = data.get("bottom", {})
    top_n_val = top_info.get("n", top_n)
    bottom_n_val = bottom_info.get("n", bottom_n)

    ws.cell(row=1, column=2, value=f"前{top_n_val}名人数").fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    ws.cell(row=1, column=2).alignment = center
    ws.cell(row=1, column=3, value=f"后{bottom_n_val}名人数").fill = header_fill
    ws.cell(row=1, column=3).font = header_font
    ws.cell(row=1, column=3).alignment = center

    # 合并所有班级
    all_classes = set()
    top_dist = top_info.get("distribution", {})
    bottom_dist = bottom_info.get("distribution", {})
    all_classes.update(top_dist.keys())
    all_classes.update(bottom_dist.keys())

    row = 2
    for cls_name in sorted(all_classes):
        ws.cell(row=row, column=1, value=cls_name).alignment = center
        ws.cell(row=row, column=2, value=top_dist.get(cls_name, 0)).alignment = center
        ws.cell(row=row, column=3, value=bottom_dist.get(cls_name, 0)).alignment = center
        row += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rank_distribution_{exam_id}.xlsx"},
    )


@router.get("/export-score-shape/{exam_id}")
def export_score_shape_excel(exam_id: int, db: Session = Depends(get_db)):
    """导出分数分布形态为Excel（含年级整体和各班级的统计表与直方图）"""
    from app.services.stats_service import export_shape_to_excel
    data = export_shape_to_excel(db, exam_id)
    if not data:
        return {"error": "无数据可导出"}
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=score_shape_{exam_id}.xlsx"},
    )


@router.get("/export-critical-students/{exam_id}")
def export_critical_students_excel(exam_id: int, db: Session = Depends(get_db)):
    """导出临界生统计为Excel"""
    data = critical_students_stats(db, exam_id)
    if not data:
        return {"error": "无数据可导出"}

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "临界生统计"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center")

    headers = ["科目", "满分", "及格线", "临界区间", "临界生人数", "临界生占比", "及格线以下", "及格线以上"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    row = 2
    subjects = data.get("subjects", {})
    for subj_name, subj_data in subjects.items():
        ws.cell(row=row, column=1, value=subj_name).alignment = center
        ws.cell(row=row, column=2, value=subj_data["full_score"]).alignment = center
        ws.cell(row=row, column=3, value=subj_data["pass_line"]).alignment = center
        ws.cell(row=row, column=4, value=f"{subj_data['lower_bound']}-{subj_data['upper_bound']}").alignment = center
        ws.cell(row=row, column=5, value=subj_data["critical_count"]).alignment = center
        ws.cell(row=row, column=6, value=f"{subj_data['critical_rate']}%").alignment = center
        ws.cell(row=row, column=7, value=subj_data["below_critical"]).alignment = center
        ws.cell(row=row, column=8, value=subj_data["above_critical"]).alignment = center
        row += 1

    # 设置列宽
    col_widths = [10, 8, 8, 12, 12, 10, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=critical_students_{exam_id}.xlsx"},
    )
