"""
统计分析服务：均分、优秀率、及格率、分数段分布、班级对比、纵向追踪
"""

import io
import traceback
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.student import Student, ClassGroup
from app.models.exam import Exam, ExamSubject
from app.models.score import Score, ConversionTemplate
from app.services.score_service import get_exam_all_totals
from app.services.grade_calc_service import calculate_grades


def _median(values: list) -> float:
    """计算中位数"""
    sorted_vals = sorted(values)
    n = len(sorted_vals)
    if n % 2 == 1:
        return sorted_vals[n // 2]
    else:
        return round((sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2, 2)


def subject_stats(db: Session, exam_subject_id: int) -> dict:
    """单科统计：均分、最高分、最低分、及格率、优秀率"""
    scores = db.query(Score).filter(Score.exam_subject_id == exam_subject_id).all()
    if not scores:
        return {}

    exam_subject = db.get(ExamSubject, exam_subject_id)
    full_score = exam_subject.full_score

    # 对赋分科目用 converted_score，其他用 raw_score
    values = []
    for s in scores:
        if exam_subject.needs_conversion and s.converted_score is not None:
            values.append(s.converted_score)
        else:
            values.append(s.raw_score)

    avg = sum(values) / len(values)
    pass_line = full_score * 0.6
    excellent_line = full_score * 0.9

    return {
        "subject": exam_subject.subject,
        "count": len(values),
        "average": round(avg, 2),
        "max": max(values),
        "min": min(values),
        "pass_rate": round(sum(1 for v in values if v >= pass_line) / len(values) * 100, 2),
        "excellent_rate": round(sum(1 for v in values if v >= excellent_line) / len(values) * 100, 2),
        "full_score": full_score,
    }


def score_distribution(db: Session, exam_subject_id: int, step: int = 10) -> dict:
    """分数段分布，返回 {labels: [...], counts: [...]}"""
    scores = db.query(Score).filter(Score.exam_subject_id == exam_subject_id).all()
    if not scores:
        return {"labels": [], "counts": []}

    exam_subject = db.get(ExamSubject, exam_subject_id)
    full_score = exam_subject.full_score

    values = []
    for s in scores:
        if exam_subject.needs_conversion and s.converted_score is not None:
            values.append(s.converted_score)
        else:
            values.append(s.raw_score)

    labels = []
    counts = []
    for low in range(0, full_score, step):
        high = low + step
        label = f"{low}-{high}"
        count = sum(1 for v in values if low <= v < high)
        labels.append(label)
        counts.append(count)
    # 最后一个区间包含满分
    if full_score % step != 0:
        count = sum(1 for v in values if v == full_score)
        if counts:
            counts[-1] += count

    return {"labels": labels, "counts": counts}


def grade_overview(db: Session, exam_id: int) -> dict:
    """年级总览：各科统计 + 总分统计"""
    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_id).all()
    subjects_stats = {}
    for es in exam_subjects:
        stats = subject_stats(db, es.id)
        if stats:
            subjects_stats[es.subject] = stats

    # 总分统计
    all_totals = get_exam_all_totals(db, exam_id)
    totals = [t["total"] for t in all_totals]
    total_stats = {}
    if totals:
        total_stats = {
            "count": len(totals),
            "average": round(sum(totals) / len(totals), 2),
            "max": max(totals),
            "min": min(totals),
        }

    return {
        "exam_id": exam_id,
        "subjects": subjects_stats,
        "total": total_stats,
    }


def combination_overview(db: Session, exam_id: int) -> dict:
    """
    按班级统计数据：各班级的考试人数、原始分总平均分、原始分最高分、
    原始分最低分、原始分中位数、赋分总平均分、赋分最高分、赋分最低分、赋分中位数。
    基于 calculate_grades() 内存计算，不写数据库。
    """
    # 查找默认赋分模板
    default_tpl = db.query(ConversionTemplate).first()

    # 始终用原始分模式获取基础数据
    try:
        result_raw = calculate_grades(db, exam_id, use_conversion=False)
    except Exception:
        traceback.print_exc()
        return {"exam_id": exam_id, "班级数据": {}, "有赋分": False, "年级汇总": {}}
    students = result_raw.get("students", [])
    if not students:
        return {"exam_id": exam_id, "班级数据": {}, "有赋分": False, "年级汇总": {}}

    # 用赋分模式计算（内存中完成赋分，不写DB）
    result_with_conv = None
    if default_tpl:
        try:
            result_with_conv = calculate_grades(db, exam_id, use_conversion=True, template_id=default_tpl.id)
        except Exception:
            result_with_conv = None

    # 检查是否有实际的赋分数据
    has_conversion = False
    if result_with_conv:
        for stu in result_with_conv.get("students", []):
            if "converted_total" in stu and stu["converted_total"] > 0:
                has_conversion = True
                break

    # 按班级分组
    class_data: dict[str, list[dict]] = {}
    for stu in students:
        cls_name = stu.get("class_name", "未知班级")
        class_data.setdefault(cls_name, []).append(stu)

    # 赋分数据映射
    conv_total_map: dict[int, float] = {}
    if has_conversion and result_with_conv:
        for stu in result_with_conv.get("students", []):
            if "converted_total" in stu:
                conv_total_map[stu["student_id"]] = stu["converted_total"]

    # 构建各班级数据
    班级数据 = {}
    全部原始分 = []
    全部赋分 = []
    for cls_name, stu_list in sorted(class_data.items()):
        # 只统计有成绩的学生（raw_scores 非空）
        has_scores_students = [s for s in stu_list if s.get("raw_scores")]
        if not has_scores_students:
            continue
        raw_totals = [s["raw_total"] for s in has_scores_students]
        全部原始分.extend(raw_totals)

        entry = {
            "考试人数": len(stu_list),
            "原始分总平均分": round(sum(raw_totals) / len(raw_totals), 2),
            "原始分最高分": max(raw_totals),
            "原始分最低分": min(raw_totals),
            "原始分中位数": _median(raw_totals),
        }

        if has_conversion:
            conv_totals = []
            for s in stu_list:
                ct = conv_total_map.get(s["student_id"])
                if ct is not None and ct > 0:
                    conv_totals.append(ct)
            if conv_totals:
                全部赋分.extend(conv_totals)
                entry["赋分总平均分"] = round(sum(conv_totals) / len(conv_totals), 2)
                entry["赋分最高分"] = max(conv_totals)
                entry["赋分最低分"] = min(conv_totals)
                entry["赋分中位数"] = _median(conv_totals)

        班级数据[cls_name] = entry

    # 年级汇总
    年级汇总 = {}
    if 全部原始分:
        年级汇总 = {
            "考试人数": len(全部原始分),
            "原始分总平均分": round(sum(全部原始分) / len(全部原始分), 2),
            "原始分最高分": max(全部原始分),
            "原始分最低分": min(全部原始分),
            "原始分中位数": _median(全部原始分),
        }
        if has_conversion and 全部赋分:
            年级汇总["赋分总平均分"] = round(sum(全部赋分) / len(全部赋分), 2)
            年级汇总["赋分最高分"] = max(全部赋分)
            年级汇总["赋分最低分"] = min(全部赋分)
            年级汇总["赋分中位数"] = _median(全部赋分)

    return {
        "exam_id": exam_id,
        "班级数据": 班级数据,
        "有赋分": has_conversion,
        "年级汇总": 年级汇总,
    }


def class_comparison(db: Session, exam_id: int) -> list[dict]:
    """班级对比：每个班级各科均分和总分均分（优化：批量查询）"""
    classes = db.query(ClassGroup).all()
    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_id).all()

    if not classes or not exam_subjects:
        return []

    # 建立学生 -> 班级映射
    students = db.query(Student).all()
    student_class_map = {s.id: s.class_id for s in students}
    class_students = {}  # class_id -> [student_ids]
    for s in students:
        class_students.setdefault(s.class_id, []).append(s.id)

    # 一次性获取所有成绩
    es_ids = [es.id for es in exam_subjects]
    all_scores = (
        db.query(Score)
        .filter(Score.exam_subject_id.in_(es_ids))
        .all()
    )

    # 按 (class_id, subject) 聚合
    class_subject_values = {}  # (class_id, subject) -> [values]
    for score in all_scores:
        cls_id = student_class_map.get(score.student_id)
        if cls_id is None:
            continue
        es = next((e for e in exam_subjects if e.id == score.exam_subject_id), None)
        if not es:
            continue
        key = (cls_id, es.subject)
        val = score.converted_score if (es.needs_conversion and score.converted_score is not None) else score.raw_score
        class_subject_values.setdefault(key, []).append(val)

    result = []
    for cls in classes:
        sids = class_students.get(cls.id, [])
        if not sids:
            continue

        cls_data = {"class_name": cls.name, "student_count": len(sids), "subjects": {}}
        for es in exam_subjects:
            values = class_subject_values.get((cls.id, es.subject), [])
            if values:
                cls_data["subjects"][es.subject] = round(sum(values) / len(values), 2)

        result.append(cls_data)

    return result


def exam_quality_analysis(db: Session, exam_id: int) -> dict:
    """
    考试质量分析表：班级×组合×学科交叉统计。
    包含最高分、最低分、班级前10均分、班级后10均分、班级均分、中位数。
    同一班级的不同组合（如物化地、物化政）分别列显示。
    基于 calculate_grades() 获取成绩数据，与考试管理保持一致。
    """
    from app.services.grade_calc_service import calculate_grades

    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_id).all()
    if not exam_subjects:
        return {"exam_id": exam_id, "classes": [], "subjects": {}}

    # 使用 calculate_grades 获取所有学生的成绩（与考试管理页面一致）
    result = calculate_grades(db, exam_id, use_conversion=False)
    students = result.get("students", [])
    if not students:
        return {"exam_id": exam_id, "classes": [], "subjects": {}}

    full_score_map = {es.subject: es.full_score for es in exam_subjects}

    # 按 班级+组合 分组
    class_combo_students: dict[tuple[str, str], list[dict]] = {}
    for stu in students:
        cls_name = stu.get("class_name", "未知")
        combo = stu.get("combination", "")
        key = (cls_name, combo)
        class_combo_students.setdefault(key, []).append(stu)

    # 排序：班级名 → 组合名
    sorted_keys = sorted(class_combo_students.keys(), key=lambda x: (x[0], x[1] or ""))

    # 班级信息（每个班级+组合一条记录）
    classes_info = []
    for key in sorted_keys:
        cls_name, combo = key
        stus = class_combo_students[key]
        classes_info.append({
            "class_name": cls_name,
            "combination": combo,
            "student_count": len(stus),
        })

    def _data_key(cls_name: str, combo: str) -> str:
        """构建 subjects.classes 的查找 key"""
        return f"{cls_name}_{combo}" if combo else cls_name

    # 计算每个学科每个班级+组合的统计
    subjects_result = {}
    for es in exam_subjects:
        subj = es.subject
        subj_classes = {}

        for key in sorted_keys:
            cls_name, combo = key
            stus = class_combo_students[key]
            values = [s["raw_scores"].get(subj) for s in stus if subj in s.get("raw_scores", {})]
            values = [v for v in values if v is not None]

            dk = _data_key(cls_name, combo)
            if not values:
                subj_classes[dk] = None
                continue

            sorted_vals = sorted(values, reverse=True)
            n = len(sorted_vals)
            top10 = sorted_vals[:min(10, n)]
            bottom10 = sorted_vals[-min(10, n):]

            subj_classes[dk] = {
                "最高分": max(values),
                "最低分": min(values),
                "班级前10名平均分": round(sum(top10) / len(top10), 2),
                "班级后10名平均分": round(sum(bottom10) / len(bottom10), 2),
                "班级均分": round(sum(values) / len(values), 2),
                "中位数": _median(values),
            }

        subjects_result[subj] = {
            "full_score": full_score_map.get(subj, 100),
            "classes": subj_classes,
        }

    return {
        "exam_id": exam_id,
        "classes": classes_info,
        "subjects": subjects_result,
    }


def score_segment_stats(db: Session, exam_id: int) -> dict:
    """
    成绩分段统计：总分50分一段，学科10分一段。
    每段显示人数和占比，学科最后一行显示合格率。
    语数英90分及格，其他学科60分及格。
    """
    from app.services.grade_calc_service import COMBINATION_SUBJECTS, ALL_SUBJECTS
    from app.services.score_service import get_exam_all_totals

    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_id).all()
    if not exam_subjects:
        return {"exam_id": exam_id, "total": {}, "subjects": {}}

    es_map = {es.subject: es for es in exam_subjects}
    es_ids = [es.id for es in exam_subjects]

    students = db.query(Student).join(ClassGroup, Student.class_id == ClassGroup.id).all()
    student_map = {s.id: s for s in students}
    all_scores = db.query(Score).filter(Score.exam_subject_id.in_(es_ids)).all() if es_ids else []

    # 按 (class_id, subject) 收集分数
    class_subject_scores: dict[tuple[int, str], list[float]] = {}
    class_student_ids: dict[int, set[int]] = {}

    for sc in all_scores:
        stu = student_map.get(sc.student_id)
        if not stu:
            continue
        es = next((e for e in exam_subjects if e.id == sc.exam_subject_id), None)
        if not es:
            continue
        val = sc.converted_score if (es.needs_conversion and sc.converted_score is not None) else sc.raw_score
        if val is not None:
            key = (stu.class_id, es.subject)
            class_subject_scores.setdefault(key, []).append(val)
            class_student_ids.setdefault(stu.class_id, set()).add(stu.id)

    if not class_student_ids:
        return {"exam_id": exam_id, "total": {}, "subjects": {}}

    # 班级名称映射
    class_name_map: dict[int, str] = {}
    for s in students:
        if s.class_id not in class_name_map:
            class_name_map[s.class_id] = s.class_group.name

    # 班级组合映射（支持多组合）
    class_combinations: dict[int, list[str]] = {}
    for cls_id, sids in class_student_ids.items():
        combos = set()
        for s in students:
            if s.id in sids and s.combination:
                combos.add(s.combination)
        class_combinations[cls_id] = sorted(combos) if combos else [""]

    # 确定每个班级的组合所属科目集合
    class_subjects_set: dict[int, set[str]] = {}
    for cls_id, combos in class_combinations.items():
        subj_set = set()
        for combo in combos:
            if combo:
                subj_set.update(COMBINATION_SUBJECTS.get(combo, ALL_SUBJECTS))
            else:
                subj_set.update(ALL_SUBJECTS)
        class_subjects_set[cls_id] = subj_set

    # 总分分段（50分一段）
    all_totals = get_exam_all_totals(db, exam_id)
    class_totals: dict[int, list[float]] = {}
    for t in all_totals:
        cls_id = t.get("class_id")
        if cls_id is not None:
            class_totals.setdefault(cls_id, []).append(t["total"])

    total_max = max((t["total"] for t in all_totals), default=0)
    total_step = 50
    total_labels = []
    for low in range(0, int(total_max) + total_step, total_step):
        high = low + total_step
        total_labels.append(f"{low}-{high}")

    total_classes = {}
    for cls_id, totals in class_totals.items():
        cls_name = class_name_map.get(cls_id, str(cls_id))
        counts = []
        for low in range(0, int(total_max) + total_step, total_step):
            high = low + total_step
            c = sum(1 for v in totals if low <= v < high)
            counts.append(c)
        # 最后一个区间包含满分
        if totals:
            max_val = max(totals)
            max_bin = int(max_val // total_step)
            if max_bin < len(counts) and max_val == max_bin + total_step:
                pass  # 已包含
        n = len(totals)
        pcts = [round(c / n * 100, 2) if n > 0 else 0 for c in counts]
        total_classes[cls_name] = {"counts": counts, "pcts": pcts}

    # 学科分段（10分一段）
    subjects_result = {}
    for es in exam_subjects:
        subj = es.subject
        full_score = es.full_score
        step = 10
        pass_threshold = 90 if subj in ("语文", "数学", "英语") else 60

        labels = []
        for low in range(0, full_score, step):
            high = low + step
            labels.append(f"{low}-{high}")
        # 满分区间
        if full_score % step != 0:
            labels.append(f"{full_score - (full_score % step)}-{full_score}")

        subj_classes = {}
        for cls_id in class_student_ids:
            cls_name = class_name_map.get(cls_id, str(cls_id))

            # 如果该班级的所有组合都不包含此科目
            if subj not in class_subjects_set.get(cls_id, set()):
                subj_classes[cls_name] = None
                continue

            values = class_subject_scores.get((cls_id, subj), [])
            if not values:
                subj_classes[cls_name] = None
                continue

            counts = []
            for low in range(0, full_score, step):
                high = low + step
                c = sum(1 for v in values if low <= v < high)
                counts.append(c)
            # 满分
            if full_score % step != 0:
                c = sum(1 for v in values if v == full_score)
                if counts:
                    counts[-1] += c

            n = len(values)
            pcts = [round(c / n * 100, 2) if n > 0 else 0 for c in counts]
            pass_count = sum(1 for v in values if v >= pass_threshold)
            pass_rate = round(pass_count / n * 100, 2) if n > 0 else 0

            subj_classes[cls_name] = {
                "counts": counts,
                "pcts": pcts,
                "pass_rate": pass_rate,
            }

        subjects_result[subj] = {
            "step": step,
            "full_score": full_score,
            "pass_threshold": pass_threshold,
            "labels": labels,
            "classes": subj_classes,
        }

    return {
        "exam_id": exam_id,
        "total": {
            "step": total_step,
            "labels": total_labels,
            "classes": total_classes,
        },
        "subjects": subjects_result,
    }


def export_quality_to_excel(db: Session, exam_id: int) -> bytes:
    """导出考试质量分析和成绩分段统计为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    quality = exam_quality_analysis(db, exam_id)
    segments = score_segment_stats(db, exam_id)

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center")
    subject_fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    subj_title_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

    # ========== Sheet 1: 考试质量分析表 ==========
    classes = quality.get("classes", [])
    subjects = quality.get("subjects", {})

    if classes and subjects:
        ws1 = wb.create_sheet(title="考试质量分析表")

        # 表头行1: 科目 | 班级 | 高一（1）班（跨组合合并）| ...
        row_idx = 1
        ws1.cell(row=row_idx, column=1, value="科目").fill = header_fill
        ws1.cell(row=row_idx, column=1).font = header_font
        ws1.cell(row=row_idx, column=1).alignment = center
        ws1.cell(row=row_idx, column=2, value="班级").fill = header_fill
        ws1.cell(row=row_idx, column=2).font = header_font
        ws1.cell(row=row_idx, column=2).alignment = center
        ci = 0
        while ci < len(classes):
            cls = classes[ci]
            span = 1
            while ci + span < len(classes) and classes[ci + span]["class_name"] == cls["class_name"]:
                span += 1
            col_start = ci + 3
            if span > 1:
                ws1.merge_cells(start_row=row_idx, start_column=col_start, end_row=row_idx, end_column=col_start + span - 1)
            cell = ws1.cell(row=row_idx, column=col_start, value=cls["class_name"])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            ci += span

        # 表头行2: | 组合 | 物化生 | 物化地 | 物化政 | ...
        row_idx = 2
        ws1.cell(row=row_idx, column=1, value="").fill = subject_fill
        ws1.cell(row=row_idx, column=2, value="组合").fill = subject_fill
        ws1.cell(row=row_idx, column=2).alignment = center
        for ci, cls in enumerate(classes):
            cell = ws1.cell(row=row_idx, column=ci + 3, value=cls["combination"] or "未分组")
            cell.fill = subject_fill
            cell.alignment = center

        # 表头行3: | 班级人数 | 50 | 50 | ...
        row_idx = 3
        ws1.cell(row=row_idx, column=1, value="").fill = subject_fill
        ws1.cell(row=row_idx, column=2, value="班级人数").fill = subject_fill
        ws1.cell(row=row_idx, column=2).alignment = center
        for ci, cls in enumerate(classes):
            cell = ws1.cell(row=row_idx, column=ci + 3, value=cls["student_count"])
            cell.fill = subject_fill
            cell.alignment = center

        def _export_data_key(cls_name, combo):
            return f"{cls_name}_{combo}" if combo else cls_name

        # 数据行
        row_idx = 4
        stat_rows = [
            ("任课教师", lambda d: ""),
            ("最高分", lambda d: d["最高分"]),
            ("最低分", lambda d: d["最低分"]),
            ("班级前10名平均分", lambda d: d["班级前10名平均分"]),
            ("班级后10名平均分", lambda d: d["班级后10名平均分"]),
            ("班级均分", lambda d: d["班级均分"]),
            ("中位数", lambda d: d.get("中位数", "")),
        ]

        for subj_name, subj_data in subjects.items():
            # 学科分隔行（与网页一致）
            full_score = subj_data.get("full_score", "")
            ws1.cell(row=row_idx, column=1, value=f"{subj_name}（满分{full_score}）").fill = subj_title_fill
            ws1.cell(row=row_idx, column=1).font = Font(bold=True)
            ws1.cell(row=row_idx, column=1).alignment = center
            for ci in range(len(classes)):
                ws1.cell(row=row_idx, column=ci + 2).fill = subj_title_fill
                ws1.cell(row=row_idx, column=ci + 3).fill = subj_title_fill
            row_idx += 1

            for stat_label, stat_fn in stat_rows:
                ws1.cell(row=row_idx, column=1, value="")
                ws1.cell(row=row_idx, column=2, value=stat_label).alignment = center
                for ci, cls in enumerate(classes):
                    dk = _export_data_key(cls["class_name"], cls.get("combination", ""))
                    cls_data = subj_data["classes"].get(dk)
                    if cls_data is None:
                        ws1.cell(row=row_idx, column=ci + 3, value="/").alignment = center
                    else:
                        val = stat_fn(cls_data)
                        ws1.cell(row=row_idx, column=ci + 3, value=val).alignment = center
                row_idx += 1

        # 班主任行
        ws1.cell(row=row_idx, column=1, value="班主任").fill = subject_fill
        ws1.cell(row=row_idx, column=1).font = Font(bold=True)
        ws1.cell(row=row_idx, column=1).alignment = center
        ws1.cell(row=row_idx, column=2, value="").fill = subject_fill
        for ci in range(len(classes)):
            ws1.cell(row=row_idx, column=ci + 3, value="").fill = subject_fill
        row_idx += 1

        # 列宽
        from openpyxl.utils import get_column_letter
        ws1.column_dimensions["A"].width = 12
        ws1.column_dimensions["B"].width = 18
        for ci in range(len(classes)):
            ws1.column_dimensions[get_column_letter(ci + 3)].width = 14

    # ========== Sheet 2: 成绩分段统计（合并为一张表） ==========
    total_data = segments.get("total", {})
    subjects_seg = segments.get("subjects", {})
    has_total = total_data and total_data.get("classes")
    has_subjects = bool(subjects_seg)

    if has_total or has_subjects:
        ws2 = wb.create_sheet(title="成绩分段统计")
        row_idx = 1

        # 获取班级列表（从总分数据或第一个学科数据中取）
        cls_names = []
        if has_total:
            cls_names = list(total_data["classes"].keys())
        elif has_subjects:
            first_subj = next(iter(subjects_seg.values()))
            cls_names = [k for k in first_subj["classes"].keys()]

        # === 总分分段 ===
        if has_total:
            labels = total_data["labels"]
            classes_data = total_data["classes"]

            # 标题行
            ws2.cell(row=row_idx, column=1, value="总分（50分一段）").fill = subj_title_fill
            ws2.cell(row=row_idx, column=1).font = Font(bold=True)
            for ci in range(len(cls_names)):
                ws2.cell(row=row_idx, column=ci + 2).fill = subj_title_fill
            row_idx += 1

            # 表头行：分数段 | 1班 | 2班 | ...
            ws2.cell(row=row_idx, column=1, value="分数段").fill = header_fill
            ws2.cell(row=row_idx, column=1).font = header_font
            ws2.cell(row=row_idx, column=1).alignment = center
            for ci, cn in enumerate(cls_names):
                cell = ws2.cell(row=row_idx, column=ci + 2, value=cn)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            row_idx += 1

            # 数据行
            for li, label in enumerate(labels):
                ws2.cell(row=row_idx, column=1, value=label).alignment = center
                for ci, cn in enumerate(cls_names):
                    cls_data = classes_data.get(cn)
                    if cls_data:
                        cnt = cls_data["counts"][li]
                        pct = cls_data["pcts"][li]
                        cell = ws2.cell(row=row_idx, column=ci + 2, value=f"{cnt}\n{pct}%")
                        cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    else:
                        cell = ws2.cell(row=row_idx, column=ci + 2, value="0\n0%")
                        cell.alignment = Alignment(horizontal="center", wrap_text=True)
                row_idx += 1

            row_idx += 1  # 空行

        # === 各学科分段 ===
        # 按顺序：语文、数学、英语、物理、化学、生物、政治、历史、地理
        subject_order = ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]
        ordered_subjects = [(s, subjects_seg[s]) for s in subject_order if s in subjects_seg]

        for subj_name, subj_data in ordered_subjects:
            labels = subj_data["labels"]
            classes_data = subj_data["classes"]
            pass_threshold = subj_data["pass_threshold"]

            # 标题行
            ws2.cell(row=row_idx, column=1, value=f"{subj_name}（{subj_data['full_score']}分满分，{pass_threshold}分及格）").fill = subj_title_fill
            ws2.cell(row=row_idx, column=1).font = Font(bold=True)
            for ci in range(len(cls_names)):
                ws2.cell(row=row_idx, column=ci + 2).fill = subj_title_fill
            row_idx += 1

            # 表头行：分数段 | 1班 | 2班 | ...
            ws2.cell(row=row_idx, column=1, value="分数段").fill = header_fill
            ws2.cell(row=row_idx, column=1).font = header_font
            ws2.cell(row=row_idx, column=1).alignment = center
            for ci, cn in enumerate(cls_names):
                cell = ws2.cell(row=row_idx, column=ci + 2, value=cn)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            row_idx += 1

            # 数据行
            for li, label in enumerate(labels):
                ws2.cell(row=row_idx, column=1, value=label).alignment = center
                for ci, cn in enumerate(cls_names):
                    cls_data = classes_data.get(cn)
                    if cls_data is None:
                        ws2.cell(row=row_idx, column=ci + 2, value="/").alignment = center
                    else:
                        cnt = cls_data["counts"][li]
                        pct = cls_data["pcts"][li]
                        cell = ws2.cell(row=row_idx, column=ci + 2, value=f"{cnt}\n{pct}%")
                        cell.alignment = Alignment(horizontal="center", wrap_text=True)
                row_idx += 1

            # 合格率行
            ws2.cell(row=row_idx, column=1, value="合格率").fill = pass_fill
            ws2.cell(row=row_idx, column=1).font = Font(bold=True)
            ws2.cell(row=row_idx, column=1).alignment = center
            for ci, cn in enumerate(cls_names):
                cls_data = classes_data.get(cn)
                if cls_data is None:
                    cell = ws2.cell(row=row_idx, column=ci + 2, value="/")
                else:
                    cell = ws2.cell(row=row_idx, column=ci + 2, value=f"{cls_data['pass_rate']}%")
                cell.fill = pass_fill
                cell.font = Font(bold=True)
                cell.alignment = center
            row_idx += 1

            row_idx += 1  # 空行

        # 列宽
        ws2.column_dimensions["A"].width = 16
        for ci in range(len(cls_names)):
            ws2.column_dimensions[ws2.cell(row=1, column=ci + 2).column_letter].width = 14

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def student_trend(db: Session, student_id: int) -> list[dict]:
    """学生个人趋势：多次考试各科成绩、总分、排名、均分对比"""
    student = db.get(Student, student_id)
    if not student:
        return []

    from app.constants import CONVERSION_SUBJECTS, COMBINATION_SUBJECTS, ALL_SUBJECTS

    exams = db.query(Exam).order_by(Exam.exam_date).all()
    trend = []

    for exam in exams:
        exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam.id).all()
        es_map = {es.subject: es for es in exam_subjects}
        es_ids = [es.id for es in exam_subjects]

        # 批量加载该考试所有成绩
        all_scores = db.query(Score).filter(Score.exam_subject_id.in_(es_ids)).all() if es_ids else []
        score_map = {}
        for sc in all_scores:
            score_map[(sc.student_id, sc.exam_subject_id)] = sc

        # 批量加载所有学生（含班级）
        all_students = db.query(Student).join(ClassGroup, Student.class_id == ClassGroup.id).all()
        student_map = {s.id: s for s in all_students}

        # 当前学生的成绩
        exam_scores = {}
        for subj, es in es_map.items():
            sc = score_map.get((student_id, es.id))
            if sc:
                if es.needs_conversion and sc.converted_score is not None:
                    exam_scores[subj] = sc.converted_score
                else:
                    exam_scores[subj] = sc.raw_score

        total = sum(exam_scores.values()) if exam_scores else 0

        # 根据考试是否在分班后，决定用哪个班级做排名
        is_post = exam.post_split

        def get_class_id(s):
            """分班前的考试用原班级，分班后的考试用现班级"""
            if is_post:
                return s.class_id
            return s.original_class_id or s.class_id

        my_class = get_class_id(student)

        # 计算每个学生的总分（用于排名）
        student_totals = []
        for s in all_students:
            subjects = COMBINATION_SUBJECTS.get(s.combination, ALL_SUBJECTS) if s.combination else ALL_SUBJECTS
            s_total = 0
            has_score = False
            for subj in subjects:
                es = es_map.get(subj)
                if es:
                    sc = score_map.get((s.id, es.id))
                    if sc:
                        val = sc.converted_score if (es.needs_conversion and sc.converted_score is not None) else sc.raw_score
                        s_total += val
                        has_score = True
            if has_score:
                student_totals.append({"id": s.id, "cid": get_class_id(s), "total": s_total})

        # 排名
        student_totals.sort(key=lambda x: x["total"], reverse=True)
        grade_rank = 0
        for i, st in enumerate(student_totals, 1):
            if st["id"] == student_id:
                grade_rank = i
                break

        # 班级排名
        class_totals = [st for st in student_totals if st["cid"] == my_class]
        class_totals.sort(key=lambda x: x["total"], reverse=True)
        class_rank = 0
        for i, st in enumerate(class_totals, 1):
            if st["id"] == student_id:
                class_rank = i
                break

        # 班级均分和年级均分（按科目）
        class_scores = {subj: [] for subj in es_map}
        grade_scores = {subj: [] for subj in es_map}
        class_total_list = []
        grade_total_list = []

        for s in all_students:
            subjects = COMBINATION_SUBJECTS.get(s.combination, ALL_SUBJECTS) if s.combination else ALL_SUBJECTS
            s_total = 0
            has_score = False
            for subj in es_map:
                es = es_map[subj]
                sc = score_map.get((s.id, es.id))
                if sc:
                    val = sc.converted_score if (es.needs_conversion and sc.converted_score is not None) else sc.raw_score
                    grade_scores[subj].append(val)
                    s_total += val
                    has_score = True
                    if get_class_id(s) == my_class:
                        class_scores[subj].append(val)
            if has_score:
                grade_total_list.append(s_total)
                if get_class_id(s) == my_class:
                    class_total_list.append(s_total)

        class_avg = {subj: round(sum(v) / len(v), 1) if v else 0 for subj, v in class_scores.items()}
        grade_avg = {subj: round(sum(v) / len(v), 1) if v else 0 for subj, v in grade_scores.items()}

        trend.append({
            "exam_name": exam.name,
            "exam_date": str(exam.exam_date),
            "scores": exam_scores,
            "total": round(total, 2),
            "class_rank": class_rank,
            "grade_rank": grade_rank,
            "class_avg": class_avg,
            "grade_avg": grade_avg,
            "class_total_avg": round(sum(class_total_list) / len(class_total_list), 1) if class_total_list else 0,
            "grade_total_avg": round(sum(grade_total_list) / len(grade_total_list), 1) if grade_total_list else 0,
        })

    return trend


def rank_distribution(db: Session, exam_id: int, top_n: int = 30, bottom_n: int = 30) -> dict:
    """
    排名分布：统计前x名和后x名中各班级学生的分布情况。
    如果考试是分班后的，则按组合内排名。
    """
    from app.services.grade_calc_service import COMBINATION_SUBJECTS, ALL_SUBJECTS
    from app.services.score_service import get_exam_all_totals

    exam = db.get(Exam, exam_id)
    if not exam:
        return {"exam_id": exam_id, "top": {}, "bottom": {}, "total_students": 0}

    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_id).all()
    if not exam_subjects:
        return {"exam_id": exam_id, "top": {}, "bottom": {}, "total_students": 0}

    es_map = {es.subject: es for es in exam_subjects}

    # 获取所有学生和班级
    students = db.query(Student).join(ClassGroup, Student.class_id == ClassGroup.id).all()
    class_name_map = {}
    for s in students:
        class_name_map[s.class_id] = s.class_group.name

    # 计算每个学生的总分
    all_totals = get_exam_all_totals(db, exam_id)
    if not all_totals:
        return {"exam_id": exam_id, "top": {}, "bottom": {}, "total_students": 0}

    # 按组合分组（如果分班后）
    is_post = exam.post_split
    if is_post:
        # 按组合分组排名
        combo_students = {}
        for t in all_totals:
            stu = next((s for s in students if s.id == t["student_id"]), None)
            if stu:
                combo = stu.combination or "未分组"
                combo_students.setdefault(combo, []).append({
                    "student_id": t["student_id"],
                    "total": t["total"],
                    "class_id": stu.class_id,
                    "class_name": class_name_map.get(stu.class_id, "")
                })

        # 对每个组合分别排名
        top_dist = {}  # {class_name: count}
        bottom_dist = {}
        total_count = 0

        for combo, stu_list in combo_students.items():
            stu_list.sort(key=lambda x: x["total"], reverse=True)
            n = len(stu_list)
            total_count += n

            # 前x名
            for s in stu_list[:top_n]:
                cls_name = s["class_name"]
                top_dist[cls_name] = top_dist.get(cls_name, 0) + 1

            # 后x名
            for s in stu_list[-bottom_n:]:
                cls_name = s["class_name"]
                bottom_dist[cls_name] = bottom_dist.get(cls_name, 0) + 1

        return {
            "exam_id": exam_id,
            "top": {"n": top_n, "distribution": top_dist},
            "bottom": {"n": bottom_n, "distribution": bottom_dist},
            "total_students": total_count,
            "by_combination": True,
        }
    else:
        # 不分班，直接按年级排名
        all_totals.sort(key=lambda x: x["total"], reverse=True)
        total_count = len(all_totals)

        top_dist = {}
        for t in all_totals[:top_n]:
            stu = next((s for s in students if s.id == t["student_id"]), None)
            if stu:
                cls_name = class_name_map.get(stu.class_id, "")
                top_dist[cls_name] = top_dist.get(cls_name, 0) + 1

        bottom_dist = {}
        for t in all_totals[-bottom_n:]:
            stu = next((s for s in students if s.id == t["student_id"]), None)
            if stu:
                cls_name = class_name_map.get(stu.class_id, "")
                bottom_dist[cls_name] = bottom_dist.get(cls_name, 0) + 1

        return {
            "exam_id": exam_id,
            "top": {"n": top_n, "distribution": top_dist},
            "bottom": {"n": bottom_n, "distribution": bottom_dist},
            "total_students": total_count,
            "by_combination": False,
        }


def _calc_skewness(values: list[float]) -> float:
    """计算偏度（Fisher's definition，与scipy一致）"""
    n = len(values)
    if n < 3:
        return 0.0
    mean = sum(values) / n
    m2 = sum((x - mean) ** 2 for x in values) / n
    m3 = sum((x - mean) ** 3 for x in values) / n
    if m2 == 0:
        return 0.0
    # 使用样本偏度公式（与scipy.stats.skew bias=True一致）
    return m3 / (m2 ** 1.5)


def _calc_kurtosis(values: list[float]) -> float:
    """计算峰度（excess kurtosis，与scipy.stats.kurtosis fisher=True一致）"""
    n = len(values)
    if n < 4:
        return 0.0
    mean = sum(values) / n
    m2 = sum((x - mean) ** 2 for x in values) / n
    m4 = sum((x - mean) ** 4 for x in values) / n
    if m2 == 0:
        return 0.0
    return m4 / (m2 ** 2) - 3.0


def _calc_histogram(values: list[float], full_score: int, bin_size: int = 10) -> dict:
    """计算直方图分箱数据"""
    if not values:
        return {"labels": [], "counts": []}
    bins = list(range(0, full_score + bin_size, bin_size))
    labels = []
    counts = []
    for i in range(len(bins) - 1):
        lo, hi = bins[i], bins[i + 1]
        cnt = sum(1 for v in values if lo <= v < hi)
        if i == len(bins) - 2:  # 最后一个区间包含上界
            cnt = sum(1 for v in values if lo <= v <= hi)
        labels.append(f"{lo}-{hi}")
        counts.append(cnt)
    return {"labels": labels, "counts": counts}


def _describe_shape(skewness: float, kurtosis: float) -> tuple[str, str, str]:
    """根据偏度和峰度返回分布形态、描述和峰度描述"""
    if abs(skewness) < 0.5:
        shape = "正态分布"
        description = "分数分布较为均匀，符合正常的考试成绩分布规律。"
    elif skewness > 0.5:
        shape = "右偏分布（正偏）"
        description = "低分段学生较多，高分段学生较少，整体成绩偏低，可能试题难度较大。"
    else:
        shape = "左偏分布（负偏）"
        description = "高分段学生较多，低分段学生较少，整体成绩偏高，可能试题难度较小。"

    if kurtosis > 1:
        peak_desc = "分布较为集中，学生成绩差异较小。"
    elif kurtosis < -1:
        peak_desc = "分布较为分散，学生成绩差异较大。"
    else:
        peak_desc = "分布形态适中。"

    return shape, description, peak_desc


def _analyze_values(values: list[float], full_score: int, bin_size: int = 10) -> dict | None:
    """对一组分数进行完整的分布形态分析"""
    if len(values) < 3:
        return None
    n = len(values)
    mean = sum(values) / n
    std = (sum((x - mean) ** 2 for x in values) / n) ** 0.5
    skewness = _calc_skewness(values)
    kurtosis = _calc_kurtosis(values)
    shape, description, peak_desc = _describe_shape(skewness, kurtosis)
    histogram = _calc_histogram(values, full_score, bin_size)
    return {
        "mean": round(mean, 2),
        "std": round(std, 2),
        "skewness": round(skewness, 3),
        "kurtosis": round(kurtosis, 3),
        "shape": shape,
        "description": description,
        "peak_description": peak_desc,
        "count": n,
        "histogram": histogram,
    }


def score_shape_analysis(db: Session, exam_id: int) -> dict:
    """
    分数分布形态分析：分析总分及各学科的分布形态（正态、偏左、偏右等）。
    返回各科目的分布形态描述、统计数据和直方图分箱数据。
    支持按班级查看分布。
    """
    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_id).all()
    if not exam_subjects:
        return {"exam_id": exam_id, "subjects": {}, "total": {}, "classes": []}

    # 加载所有学生和班级
    students = db.query(Student).join(ClassGroup, Student.class_id == ClassGroup.id).all()
    student_map = {s.id: s for s in students}
    class_groups = sorted(
        {(s.class_id, s.class_group.name) for s in students},
        key=lambda x: x[1]
    )
    classes_list = [{"id": cid, "name": cname} for cid, cname in class_groups]

    results = {"subjects": {}, "total": {}, "classes": classes_list}

    # 分析各科目
    for es in exam_subjects:
        scores = db.query(Score).filter(Score.exam_subject_id == es.id).all()
        if not scores:
            continue

        # 年级整体
        grade_values = []
        class_values_map: dict[int, list[float]] = {}
        for s in scores:
            val = s.converted_score if (es.needs_conversion and s.converted_score is not None) else s.raw_score
            if val is not None:
                grade_values.append(val)
                stu = student_map.get(s.student_id)
                if stu:
                    class_values_map.setdefault(stu.class_id, []).append(val)

        grade_result = _analyze_values(grade_values, es.full_score)
        if not grade_result:
            continue

        # 按班级分析
        per_class = {}
        for cid, cname in class_groups:
            cv = class_values_map.get(cid, [])
            cr = _analyze_values(cv, es.full_score)
            if cr:
                per_class[cid] = cr

        results["subjects"][es.subject] = {
            **grade_result,
            "full_score": es.full_score,
            "per_class": per_class,
        }

    # 分析总分
    all_totals = get_exam_all_totals(db, exam_id)
    if all_totals:
        grade_totals = []
        class_totals_map: dict[int, list[float]] = {}
        for t in all_totals:
            if t["total"] > 0:
                grade_totals.append(t["total"])
                stu = student_map.get(t["student_id"])
                if stu:
                    class_totals_map.setdefault(stu.class_id, []).append(t["total"])

        total_full_score = sum(es.full_score for es in exam_subjects)
        grade_total_result = _analyze_values(grade_totals, total_full_score, bin_size=50)
        if grade_total_result:
            per_class_total = {}
            for cid, cname in class_groups:
                cv = class_totals_map.get(cid, [])
                cr = _analyze_values(cv, total_full_score, bin_size=50)
                if cr:
                    per_class_total[cid] = cr
            results["total"] = {**grade_total_result, "per_class": per_class_total}

    results["exam_id"] = exam_id
    return results


def export_shape_to_excel(db: Session, exam_id: int) -> bytes:
    """导出分数分布形态为Excel，包含年级整体和各班级的统计表与直方图"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter

    data = score_shape_analysis(db, exam_id)
    if not data:
        return None

    subjects = data.get("subjects", {})
    total = data.get("total", {})
    classes = data.get("classes", [])

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=13, color="1F4E79")
    section_font = Font(bold=True, size=11, color="2E75B6")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    subj_fill = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")

    summary_headers = ["科目", "满分", "参考人数", "平均分", "标准差", "偏度", "峰度", "分布形态", "分布说明"]
    col_widths = [10, 8, 10, 10, 10, 10, 10, 16, 42]
    chart_colors = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47",
                    "264478", "9B59B6", "E74C3C"]

    def build_sheet(ws, sheet_title, info_map, total_info):
        """向工作表写入汇总表 + 直方图"""
        ws.merge_cells("A1:I1")
        c = ws.cell(row=1, column=1, value=sheet_title)
        c.font = title_font
        c.alignment = Alignment(horizontal="center", vertical="center")

        # ---- 汇总表 ----
        row = 3
        for i, h in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
        row += 1

        if info_map:
            for subj_name, subj_data in info_map.items():
                fs = subj_data.get("full_score", "")
                ws.cell(row=row, column=1, value=subj_name).alignment = center
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2, value=fs).alignment = center
                ws.cell(row=row, column=2).border = thin_border
                for ci, key in enumerate(["count", "mean", "std", "skewness", "kurtosis", "shape"], 3):
                    cell = ws.cell(row=row, column=ci, value=subj_data.get(key, ""))
                    cell.alignment = center
                    cell.border = thin_border
                cell = ws.cell(row=row, column=9, value=subj_data.get("description", ""))
                cell.border = thin_border
                row += 1

        if total_info and total_info.get("shape"):
            row += 1
            for ci, key in enumerate(["", "", "count", "mean", "std", "skewness", "kurtosis", "shape"], 1):
                if ci == 1:
                    cell = ws.cell(row=row, column=1, value="总分")
                    cell.font = Font(bold=True)
                elif ci == 2:
                    cell = ws.cell(row=row, column=2, value="-")
                else:
                    cell = ws.cell(row=row, column=ci, value=total_info.get(key, ""))
                cell.alignment = center
                cell.border = thin_border
                cell.fill = subj_fill
            cell = ws.cell(row=row, column=9, value=total_info.get("description", ""))
            cell.border = thin_border
            cell.fill = subj_fill
            row += 1

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ---- 直方图区 ----
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value="▼ 直方图分布")
        c.font = section_font
        row += 2

        chart_entries = []
        if info_map:
            for subj_name, subj_data in info_map.items():
                hist = subj_data.get("histogram")
                if hist and hist.get("labels"):
                    chart_entries.append({
                        "name": f"{subj_name}（满分{subj_data.get('full_score', '')}）",
                        "subtitle": f"{subj_data.get('shape', '')} | 均分{subj_data.get('mean', '')} | 偏度{subj_data.get('skewness', '')}",
                        "histogram": hist,
                        "color": chart_colors[len(chart_entries) % len(chart_colors)],
                    })
        if total_info and total_info.get("histogram") and total_info["histogram"].get("labels"):
            chart_entries.append({
                "name": "总分分布",
                "subtitle": f"{total_info.get('shape', '')} | 均分{total_info.get('mean', '')} | 偏度{total_info.get('skewness', '')}",
                "histogram": total_info["histogram"],
                "color": "C00000",
            })

        if not chart_entries:
            return

        chart_start_row = row
        left_row = chart_start_row
        right_row = chart_start_row
        left_col_data = 1
        right_col_data = 6

        for idx, entry in enumerate(chart_entries):
            is_left = (idx % 2 == 0)
            data_col = left_col_data if is_left else right_col_data
            anchor_col = data_col + 2
            current_row = left_row if is_left else right_row

            hist = entry["histogram"]
            labels = hist["labels"]
            counts = hist["counts"]
            n_bins = len(labels)

            # 图表标题
            ws.merge_cells(start_row=current_row, start_column=data_col,
                           end_row=current_row, end_column=data_col + 1)
            c = ws.cell(row=current_row, column=data_col, value=entry["name"])
            c.font = Font(bold=True, size=9, color=entry["color"])
            current_row += 1

            # hist 表头
            ws.cell(row=current_row, column=data_col, value="分数段").font = Font(bold=True, size=8)
            ws.cell(row=current_row, column=data_col + 1, value="人数").font = Font(bold=True, size=8)
            current_row += 1

            # hist 数据
            data_start_row = current_row
            for i in range(n_bins):
                ws.cell(row=current_row, column=data_col, value=labels[i]).alignment = center
                ws.cell(row=current_row, column=data_col + 1, value=counts[i]).alignment = center
                current_row += 1
            data_end_row = current_row - 1

            # 创建图表
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = entry["name"]
            chart.y_axis.title = "人数"
            chart.x_axis.title = "分数段"
            chart.width = 18
            chart.height = 11

            data_ref = Reference(ws, min_col=data_col + 1, min_row=data_start_row - 1,
                                 max_col=data_col + 1, max_row=data_end_row)
            cats_ref = Reference(ws, min_col=data_col, min_row=data_start_row,
                                 max_col=data_col, max_row=data_end_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None

            s = chart.series[0]
            s.graphicalProperties.solidFill = entry["color"]
            s.graphicalProperties.line.solidFill = entry["color"]
            try:
                s.dLbls = DataLabelList()
                s.dLbls.showVal = True
                s.dLbls.numFmt = "0"
            except Exception:
                pass

            chart_anchor = f"{get_column_letter(anchor_col)}{data_start_row - 1}"
            ws.add_chart(chart, chart_anchor)

            current_row += 1
            if is_left:
                left_row = current_row
            else:
                right_row = current_row

    # ========== 年级整体 ==========
    ws_grade = wb.create_sheet(title="年级整体")
    build_sheet(ws_grade, "年级整体 · 分数分布形态分析", subjects, total)

    # ========== 各班级 ==========
    if classes:
        class_name_map = {c["id"]: c["name"] for c in classes}
        all_class_ids = set()
        for subj_data in subjects.values():
            for cid in (subj_data.get("per_class") or {}).keys():
                all_class_ids.add(cid)
        for cid in (total.get("per_class") or {}).keys():
            all_class_ids.add(cid)

        for cid in sorted(all_class_ids):
            cname = class_name_map.get(cid, f"班级{cid}")
            sheet_name = cname[:31]
            ws_cls = wb.create_sheet(title=sheet_name)

            cls_subjects = {}
            for subj_name, subj_data in subjects.items():
                pc = subj_data.get("per_class", {}).get(cid)
                if pc:
                    cls_subjects[subj_name] = {**pc, "full_score": subj_data.get("full_score", "")}

            cls_total = (total.get("per_class") or {}).get(cid)
            build_sheet(ws_cls, f"{cname} · 分数分布形态分析", cls_subjects, cls_total)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def critical_students_stats(db: Session, exam_id: int) -> dict:
    """
    临界生统计：统计各学科临界生人数。
    临界生定义：分数在及格线附近（及格线-10分到及格线+10分）的学生。
    """
    from app.services.grade_calc_service import COMBINATION_SUBJECTS, ALL_SUBJECTS

    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_id).all()
    if not exam_subjects:
        return {"exam_id": exam_id, "subjects": {}}

    results = {"subjects": {}}

    for es in exam_subjects:
        scores = db.query(Score).filter(Score.exam_subject_id == es.id).all()
        if not scores:
            continue

        values = []
        for s in scores:
            if es.needs_conversion and s.converted_score is not None:
                values.append(s.converted_score)
            else:
                values.append(s.raw_score)

        if not values:
            continue

        # 及格线
        pass_line = 90 if es.subject in ("语文", "数学", "英语") else 60
        # 临界区间：及格线-10分到及格线+10分
        lower_bound = pass_line - 10
        upper_bound = pass_line + 10

        critical_count = sum(1 for v in values if lower_bound <= v <= upper_bound)
        below_critical = sum(1 for v in values if v < lower_bound)
        above_critical = sum(1 for v in values if v > upper_bound)
        total_count = len(values)

        results["subjects"][es.subject] = {
            "full_score": es.full_score,
            "pass_line": pass_line,
            "lower_bound": lower_bound,
            "upper_bound": upper_bound,
            "critical_count": critical_count,
            "below_critical": below_critical,
            "above_critical": above_critical,
            "total_count": total_count,
            "critical_rate": round(critical_count / total_count * 100, 2) if total_count > 0 else 0,
        }

    results["exam_id"] = exam_id
    return results


def detect_weak_subjects(db: Session, student_id: int) -> list[dict]:
    """检测薄弱科目：最近一次考试中低于班级均分10分以上的科目"""
    student = db.get(Student, student_id)
    if not student:
        return []

    from app.constants import CONVERSION_SUBJECTS, COMBINATION_SUBJECTS, ALL_SUBJECTS

    latest_exam = db.query(Exam).order_by(Exam.exam_date.desc(), Exam.id.desc()).first()
    if not latest_exam:
        return []

    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == latest_exam.id).all()
    es_map = {es.subject: es for es in exam_subjects}
    es_ids = [es.id for es in exam_subjects]

    all_scores = db.query(Score).filter(Score.exam_subject_id.in_(es_ids)).all() if es_ids else []
    score_map = {}
    for sc in all_scores:
        score_map[(sc.student_id, sc.exam_subject_id)] = sc

    all_students = db.query(Student).join(ClassGroup, Student.class_id == ClassGroup.id).all()

    # 根据考试是否在分班后决定班级
    is_post = latest_exam.post_split
    my_class = student.class_id if is_post else (student.original_class_id or student.class_id)

    weak = []
    for subj, es in es_map.items():
        # 学生成绩
        sc = score_map.get((student_id, es.id))
        if not sc:
            continue
        student_val = sc.converted_score if (es.needs_conversion and sc.converted_score is not None) else sc.raw_score

        # 班级均分
        class_vals = []
        for s in all_students:
            s_class = s.class_id if is_post else (s.original_class_id or s.class_id)
            if s_class != my_class:
                continue
            sc2 = score_map.get((s.id, es.id))
            if sc2:
                val = sc2.converted_score if (es.needs_conversion and sc2.converted_score is not None) else sc2.raw_score
                class_vals.append(val)

        if not class_vals:
            continue

        class_avg = sum(class_vals) / len(class_vals)
        diff = student_val - class_avg
        if diff < -10:
            weak.append({
                "subject": subj,
                "score": student_val,
                "class_avg": round(class_avg, 1),
                "diff": round(diff, 1),
            })

    return sorted(weak, key=lambda x: x["diff"])
