"""
成绩计算服务

包含赋分模板管理和成绩计算两大功能。
赋分计算基于用户自定义模板（等级占比），在内存中完成，不写入数据库。
"""

import json
import math
from dataclasses import dataclass

import io
import pandas as pd
from sqlalchemy.orm import Session

from app.models.score import Score, ConversionTemplate
from app.models.exam import Exam, ExamSubject
from app.models.student import Student, ClassGroup
from app.constants import COMBINATION_SUBJECTS, ALL_SUBJECTS, CONVERSION_SUBJECTS, COMBINATION_ELECTIVES, ELECTIVE_SUBJECTS

# 各等级对应的赋分区间（固定，按江苏省标准）
TIER_SCORE_RANGES = {
    "A": (86, 100),
    "B": (71, 85),
    "C": (56, 70),
    "D": (41, 55),
    "E": (30, 40),
}

# 各等级的子等级数量
TIER_SUB_COUNTS = {"A": 5, "B": 6, "C": 5, "D": 3, "E": 2}


def auto_detect_combinations(students, es_map, score_map):
    """
    根据学生有成绩的选考科目，自动推断选科组合。
    如果学生已有 combination 则跳过。
    """
    for stu in students:
        if stu.combination:
            continue
        has_subjects = set()
        for subj in ELECTIVE_SUBJECTS:
            if subj in es_map:
                sc = score_map.get((stu.id, es_map[subj].id))
                if sc and sc.raw_score is not None:
                    has_subjects.add(subj)
        if len(has_subjects) == 3:
            for combo_name, combo_elec in COMBINATION_ELECTIVES.items():
                if has_subjects == combo_elec:
                    stu.combination = combo_name
                    break


@dataclass
class SubTierRule:
    """子等级规则"""
    tier: str
    percentile_low: float
    percentile_high: float
    converted_low: int
    converted_high: int


# ==================== 模板 CRUD ====================


def create_template(db: Session, name: str, tier_ratios: dict) -> ConversionTemplate:
    tpl = ConversionTemplate(name=name, tier_ratios=json.dumps(tier_ratios, ensure_ascii=False))
    db.add(tpl)
    db.commit()
    db.refresh(tpl)
    return tpl


def list_templates(db: Session) -> list[ConversionTemplate]:
    return db.query(ConversionTemplate).order_by(ConversionTemplate.id).all()


def get_template(db: Session, template_id: int) -> ConversionTemplate | None:
    return db.get(ConversionTemplate, template_id)


def update_template(db: Session, template_id: int, name: str | None, tier_ratios: dict | None) -> ConversionTemplate:
    tpl = db.get(ConversionTemplate, template_id)
    if not tpl:
        raise ValueError("模板不存在")
    if name is not None:
        tpl.name = name
    if tier_ratios is not None:
        tpl.tier_ratios = json.dumps(tier_ratios, ensure_ascii=False)
    db.commit()
    db.refresh(tpl)
    return tpl


def delete_template(db: Session, template_id: int) -> None:
    tpl = db.get(ConversionTemplate, template_id)
    if not tpl:
        raise ValueError("模板不存在")
    db.delete(tpl)
    db.commit()


# ==================== 赋分规则构建 ====================


def build_conversion_rules(tier_ratios: dict, total_count: int) -> tuple[list[SubTierRule], dict[str, int]]:
    """
    根据用户定义的5个等级占比和实际考生人数，推算出21个子等级的百分位区间和赋分区间。
    赋分区间固定（江苏省标准），百分位区间按实际人数计算。
    核心思路：使用"最大余额法"（Hamilton方法）分配各等级人数，
    确保四舍五入后的占比等于目标占比。
    返回 (rules, tier_counts)。
    """
    tiers = ["A", "B", "C", "D", "E"]

    # 确保占比总和为100%（归一化处理）
    total_ratio = sum(float(tier_ratios.get(t, 0)) for t in tiers)
    if total_ratio > 0 and abs(total_ratio - 100) > 0.01:
        tier_ratios = {t: float(tier_ratios.get(t, 0)) * 100 / total_ratio for t in tiers}

    # 计算每个等级的精确人数
    exact_counts = {t: total_count * float(tier_ratios.get(t, 0)) / 100.0 for t in tiers}

    # 最大余额法：先取 floor，再将余额按小数部分从大到小分配
    tier_counts = {t: math.floor(v) for t, v in exact_counts.items()}
    remainder = total_count - sum(tier_counts.values())
    sorted_by_frac = sorted(
        tiers,
        key=lambda t: exact_counts[t] - math.floor(exact_counts[t]),
        reverse=True,
    )
    for t in sorted_by_frac:
        if remainder <= 0:
            break
        tier_counts[t] += 1
        remainder -= 1

    # 根据人数计算百分位边界
    boundaries = {}  # tier -> (p_low, p_high)
    position = 0  # 当前学生位置（从0开始）
    for t in tiers:
        count = tier_counts[t]
        p_low = (position / total_count) * 100
        position += count
        p_high = (position / total_count) * 100
        boundaries[t] = (p_low, p_high)

    # 构建子等级规则
    rules = []
    for broad_tier in tiers:
        p_low, p_high = boundaries[broad_tier]
        n_sub = TIER_SUB_COUNTS[broad_tier]
        score_low, score_high = TIER_SCORE_RANGES[broad_tier]
        p_step = (p_high - p_low) / n_sub
        s_step = (score_high - score_low) / n_sub

        for i in range(n_sub):
            sub_p_low = p_low + i * p_step
            sub_p_high = p_low + (i + 1) * p_step
            # 子等级编号从高到低：A1最高，A5最低
            sub_tier = f"{broad_tier}{i + 1}"
            # 赋分也是从高到低分配
            sub_s_high = score_high - i * s_step
            sub_s_low = score_high - (i + 1) * s_step
            rules.append(SubTierRule(
                tier=sub_tier,
                percentile_low=sub_p_low,
                percentile_high=sub_p_high,
                converted_low=round(sub_s_low),
                converted_high=round(sub_s_high),
            ))

    return rules, tier_counts


def build_per_subject_rules(
    all_ratios: dict, per_subject_counts: dict[str, int]
) -> tuple[dict[str, list[SubTierRule]], dict[str, dict[str, int]]]:
    """
    根据按科目分组的等级占比和各科目考生人数，为每个赋分科目构建赋分规则。
    返回 (per_subject_rules, per_subject_tier_counts)。
    """
    result_rules = {}
    result_counts = {}
    for subj in CONVERSION_SUBJECTS:
        ratios = all_ratios.get(subj, {})
        count = per_subject_counts.get(subj, 0)
        if ratios and count > 0:
            rules, tier_counts = build_conversion_rules(ratios, count)
            result_rules[subj] = rules
            result_counts[subj] = tier_counts
    return result_rules, result_counts


def _find_tier_by_rank(rank: int, tier_boundaries: dict[str, int]) -> str:
    """根据排名位置找到对应的大等级"""
    for tier, boundary in tier_boundaries.items():
        if rank < boundary:
            return tier
    return list(tier_boundaries.keys())[-1]


# ==================== 赋分计算（内存中，不写DB） ====================


def convert_subject_with_rules(
    db: Session, exam_subject_id: int, rules: list[SubTierRule], tier_counts: dict[str, int]
) -> tuple[dict[int, dict], list[dict]]:
    """
    对某科目按自定义规则进行赋分计算（纯内存，不修改数据库）。
    使用基于排名的等级分配，确保相同原始分的学生一定在同一等级。
    返回 (学生赋分结果, 各等级统计)。
    """
    scores = (
        db.query(Score)
        .filter(Score.exam_subject_id == exam_subject_id)
        .order_by(Score.raw_score.desc())
        .all()
    )
    if not scores:
        return {}, []

    total = len(scores)
    tier_display_order = ["A", "B", "C", "D", "E"]

    # 使用传入的 tier_counts 直接构建排名边界
    tier_boundaries: dict[str, int] = {}
    cumulative = 0
    for broad_tier in tier_display_order:
        cumulative += tier_counts.get(broad_tier, 0)
        tier_boundaries[broad_tier] = cumulative

    # 按排名分档，相同原始分的学生使用相同的排名（最优排名），保证在同一等级
    # 先按分组处理相同分数
    score_rank: dict[int, int] = {}  # raw_score -> best rank_idx
    for rank_idx, score in enumerate(scores):
        if score.raw_score not in score_rank:
            score_rank[score.raw_score] = rank_idx

    # 分档：相同原始分 → 相同等级
    tier_map: dict[str, list[int]] = {}
    for rank_idx, score in enumerate(scores):
        best_rank = score_rank[score.raw_score]
        tier = _find_tier_by_rank(best_rank, tier_boundaries)
        score._tier = tier
        tier_map.setdefault(tier, []).append(rank_idx)

    # 档内线性插值
    result: dict[int, dict] = {}
    for tier_name, indices in tier_map.items():
        tier_sub_rules = [r for r in rules if r.tier[0] == tier_name]
        if not tier_sub_rules:
            continue
        t_low = tier_sub_rules[0].converted_low
        t_high = tier_sub_rules[-1].converted_high

        tier_students = [scores[i] for i in indices]
        y_high = max(s.raw_score for s in tier_students)
        y_low = min(s.raw_score for s in tier_students)

        for s in tier_students:
            if y_high == y_low:
                converted = (t_low + t_high) / 2
            else:
                converted = t_low + (s.raw_score - y_low) * (t_high - t_low) / (y_high - y_low)
            converted = round(converted)
            converted = max(30, min(100, converted))
            result[s.student_id] = {
                "tier": tier_name[0],  # 只返回大等级字母（A/B/C/D/E），不带子等级编号
                "converted": float(converted),
                "raw": s.raw_score,
            }

    # 按赋分排名
    ranked = sorted(result.items(), key=lambda x: x[1]["converted"], reverse=True)
    for rank, (sid, info) in enumerate(ranked, 1):
        info["rank"] = rank

    # 构建各等级统计报告
    tier_stats = []
    tier_display_order = ["A", "B", "C", "D", "E"]
    for broad_tier in tier_display_order:
        # 收集属于该大等级的所有学生（子等级 A1, A2, ... 归为 A）
        tier_students_info = [
            info for info in result.values() if info["tier"] == broad_tier
        ]
        if not tier_students_info:
            # 没有学生在此等级，仍然显示（人数为0）
            score_range = TIER_SCORE_RANGES[broad_tier]
            tier_stats.append({
                "tier": broad_tier,
                "ratio": 0.0,
                "count": 0,
                "raw_max": 0,
                "raw_min": 0,
                "score_range": f"{score_range[0]}~{score_range[1]}",
                "converted_avg": 0.0,
            })
            continue

        count = len(tier_students_info)
        ratio = round(count / total * 100, 2)
        raw_max = max(info["raw"] for info in tier_students_info)
        raw_min = min(info["raw"] for info in tier_students_info)
        score_range = TIER_SCORE_RANGES[broad_tier]
        converted_avg = round(
            sum(info["converted"] for info in tier_students_info) / count, 1
        )

        tier_stats.append({
            "tier": broad_tier,
            "ratio": ratio,
            "count": count,
            "raw_max": raw_max,
            "raw_min": raw_min,
            "score_range": f"{score_range[0]}~{score_range[1]}",
            "converted_avg": converted_avg,
        })

    return result, tier_stats


# ==================== 成绩计算主函数 ====================


def calculate_grades(
    db: Session,
    exam_id: int,
    use_conversion: bool = False,
    template_id: int | None = None,
) -> dict:
    """
    成绩计算主函数。
    返回结构化的计算结果，包含原始总分、班级排名、组合排名，
    以及（可选的）赋分信息。
    """
    exam = db.get(Exam, exam_id)
    if not exam:
        raise ValueError("考试不存在")

    # 批量加载数据
    students = db.query(Student).join(ClassGroup, Student.class_id == ClassGroup.id).all()
    if not students:
        return {"exam_id": exam_id, "exam_name": exam.name, "use_conversion": use_conversion, "students": []}

    is_post = exam.post_split

    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_id).all()
    es_map = {es.subject: es for es in exam_subjects}
    es_ids = [es.id for es in exam_subjects]

    all_scores = db.query(Score).filter(Score.exam_subject_id.in_(es_ids)).all() if es_ids else []
    score_map = {(s.student_id, s.exam_subject_id): s for s in all_scores}

    # 自动推断选科组合
    auto_detect_combinations(students, es_map, score_map)

    # 赋分计算（如果启用）
    conversion_results: dict[str, dict[int, dict]] = {}  # subject -> {student_id: info}
    conversion_report: dict[str, list[dict]] = {}  # subject -> 各等级统计
    template_info = None
    if use_conversion and template_id:
        tpl = db.get(ConversionTemplate, template_id)
        if not tpl:
            raise ValueError("赋分模板不存在")
        template_info = {"id": tpl.id, "name": tpl.name, "tier_ratios": tpl.ratios_dict}
        # 统计每个赋分科目的考生人数
        per_subject_counts: dict[str, int] = {}
        for es in exam_subjects:
            if es.subject in CONVERSION_SUBJECTS:
                count = sum(1 for s in all_scores if s.exam_subject_id == es.id and s.raw_score is not None)
                per_subject_counts[es.subject] = count
        per_subject_rules, per_subject_tier_counts = build_per_subject_rules(tpl.ratios_dict, per_subject_counts)

        for es in exam_subjects:
            if es.subject in CONVERSION_SUBJECTS and es.subject in per_subject_rules:
                conv, tier_stats = convert_subject_with_rules(
                    db, es.id, per_subject_rules[es.subject], per_subject_tier_counts[es.subject]
                )
                conversion_results[es.subject] = conv
                if tier_stats:
                    conversion_report[es.subject] = tier_stats

    # 构建学生数据
    student_data = []
    for stu in students:
        combination = stu.combination or ""
        subjects = COMBINATION_SUBJECTS.get(combination, ALL_SUBJECTS) if combination else ALL_SUBJECTS

        # 收集所有9科原始分
        raw_scores = {}
        for subj in ALL_SUBJECTS:
            if subj in es_map:
                sc = score_map.get((stu.id, es_map[subj].id))
                if sc:
                    raw_scores[subj] = sc.raw_score

        # 原始总分（按组合的6科）
        raw_total = 0.0
        has_score = False
        for subj in subjects:
            if subj in raw_scores:
                raw_total += raw_scores[subj]
                has_score = True
        raw_total = round(raw_total, 1) if has_score else 0.0

        # 分班前的考试用原班级，分班后的考试用现班级
        ranking_cid = stu.class_id if is_post else (stu.original_class_id or stu.class_id)
        ranking_cname = stu.class_group.name if is_post else (
            stu.original_class_group.name if stu.original_class_group else stu.class_group.name
        )

        entry = {
            "student_id": stu.id,
            "student_no": stu.student_no,
            "student_name": stu.name,
            "class_name": ranking_cname,
            "class_id": ranking_cid,
            "combination": combination,
            "raw_scores": raw_scores,
            "raw_total": raw_total,
        }

        # 赋分信息
        if use_conversion and conversion_results:
            conversion_details = {}
            converted_total = raw_total
            for subj in CONVERSION_SUBJECTS:
                if subj in conversion_results and stu.id in conversion_results[subj]:
                    info = conversion_results[subj][stu.id]
                    conversion_details[subj] = {
                        "raw": info["raw"],
                        "tier": info["tier"],
                        "converted": info["converted"],
                        "rank": info["rank"],
                    }
                    # 如果该科目在学生组合中，用赋分替换原始分计算总分
                    if subj in subjects and subj in raw_scores:
                        converted_total = converted_total - raw_scores[subj] + info["converted"]
            entry["conversion_details"] = conversion_details
            entry["converted_total"] = round(converted_total, 1)

        student_data.append(entry)

    # 排名计算
    _assign_rankings(student_data, use_conversion)

    return {
        "exam_id": exam_id,
        "exam_name": exam.name,
        "use_conversion": use_conversion,
        "template": template_info,
        "students": student_data,
        "conversion_report": conversion_report,
    }


def _assign_rankings(student_data: list[dict], use_conversion: bool) -> None:
    """为每个学生分配班级排名和组合排名"""
    # 班级排名（按原始总分）
    class_groups: dict[int, list[dict]] = {}
    for s in student_data:
        class_groups.setdefault(s["class_id"], []).append(s)
    for group in class_groups.values():
        group.sort(key=lambda x: x["raw_total"], reverse=True)
        for rank, s in enumerate(group, 1):
            s["class_rank"] = rank

    # 组合排名（按原始总分）：空组合的学生参与全部排名
    combo_groups: dict[str, list[dict]] = {}
    for s in student_data:
        combo = s["combination"] if s["combination"] else "__ALL__"
        combo_groups.setdefault(combo, []).append(s)
    for group in combo_groups.values():
        group.sort(key=lambda x: x["raw_total"], reverse=True)
        for rank, s in enumerate(group, 1):
            s["combination_rank"] = rank

    # 赋分排名
    if use_conversion:
        # 赋分班级排名
        class_groups_conv: dict[int, list[dict]] = {}
        for s in student_data:
            class_groups_conv.setdefault(s["class_id"], []).append(s)
        for group in class_groups_conv.values():
            group.sort(key=lambda x: x.get("converted_total", 0), reverse=True)
            for rank, s in enumerate(group, 1):
                s["converted_class_rank"] = rank

        # 赋分组合排名
        combo_groups_conv: dict[str, list[dict]] = {}
        for s in student_data:
            combo = s["combination"] if s["combination"] else "__ALL__"
            combo_groups_conv.setdefault(combo, []).append(s)
        for group in combo_groups_conv.values():
            group.sort(key=lambda x: x.get("converted_total", 0), reverse=True)
            for rank, s in enumerate(group, 1):
                s["converted_combination_rank"] = rank


# ==================== 导出 ====================


def _get_export_columns(use_conversion: bool) -> list[str]:
    """构建导出列顺序，与页面 renderTable 完全一致"""
    columns = ["学号", "姓名", "班级", "组合"]
    columns += ALL_SUBJECTS
    columns += ["原始总分", "班级排名", "组合排名"]
    if use_conversion:
        for subj in CONVERSION_SUBJECTS:
            columns += [f"{subj}等级", f"{subj}赋分", f"{subj}赋分排名"]
        columns += ["赋分总分", "赋分班级排名", "赋分组合排名"]
    return columns


def _build_student_rows(students: list[dict], use_conversion: bool, columns: list[str]) -> list[dict]:
    """将学生数据转为行字典列表"""
    rows = []
    for s in students:
        row = {
            "学号": s["student_no"],
            "姓名": s["student_name"],
            "班级": s["class_name"],
            "组合": s["combination"],
        }
        for subj in ALL_SUBJECTS:
            row[subj] = s["raw_scores"].get(subj, "")
        row["原始总分"] = s["raw_total"]
        row["班级排名"] = s.get("class_rank", "")
        row["组合排名"] = s.get("combination_rank", "")

        if use_conversion and "conversion_details" in s:
            for subj in CONVERSION_SUBJECTS:
                if subj in s["conversion_details"]:
                    d = s["conversion_details"][subj]
                    row[f"{subj}等级"] = d["tier"]
                    row[f"{subj}赋分"] = d["converted"]
                    row[f"{subj}赋分排名"] = d["rank"]
                else:
                    row[f"{subj}等级"] = ""
                    row[f"{subj}赋分"] = ""
                    row[f"{subj}赋分排名"] = ""
            row["赋分总分"] = s.get("converted_total", "")
            row["赋分班级排名"] = s.get("converted_class_rank", "")
            row["赋分组合排名"] = s.get("converted_combination_rank", "")

        rows.append(row)
    return rows


def _apply_sheet_styling(ws, columns: list[str], rows: list[dict], use_conversion: bool):
    """为工作表设置表头样式、列底色和列宽"""
    from openpyxl.styles import PatternFill, Font, Alignment

    fill_primary = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
    fill_info = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fill_success = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    raw_total_cols = {"原始总分", "班级排名", "组合排名"}
    conv_detail_cols = set()
    conv_total_cols = {"赋分总分", "赋分班级排名", "赋分组合排名"}
    if use_conversion:
        for subj in CONVERSION_SUBJECTS:
            conv_detail_cols.add(f"{subj}等级")
            conv_detail_cols.add(f"{subj}赋分")
            conv_detail_cols.add(f"{subj}赋分排名")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(2, len(rows) + 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center")
            if col_name in raw_total_cols:
                cell.fill = fill_primary
            elif col_name in conv_detail_cols:
                cell.fill = fill_info
            elif col_name in conv_total_cols:
                cell.fill = fill_success

    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 20)


def _sanitize_sheet_name(name: str) -> str:
    """清理工作表名称，去除 Excel 不支持的字符，截断到31字符"""
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, '_')
    return name[:31]


def export_grade_results(
    db: Session,
    exam_id: int,
    use_conversion: bool = False,
    template_id: int | None = None,
    filter_type: str = "all",
    filter_value: str | list | None = None,
) -> bytes:
    """
    导出成绩计算结果为 Excel，列顺序与页面计算结果一致，带底色区分。
    分班级/分组合导出时，每个班级/组合单独一张工作表。
    """
    from openpyxl import Workbook

    result = calculate_grades(db, exam_id, use_conversion, template_id)
    all_students = result["students"]
    if not all_students:
        return b""

    columns = _get_export_columns(use_conversion)

    # 分组：决定工作表结构
    if filter_type == "class" and filter_value:
        selected = filter_value if isinstance(filter_value, list) else [filter_value]
        groups: dict[str, list] = {}
        for s in all_students:
            if s["class_name"] in selected:
                groups.setdefault(s["class_name"], []).append(s)
    elif filter_type == "combination" and filter_value:
        selected = filter_value if isinstance(filter_value, list) else [filter_value]
        groups = {}
        for s in all_students:
            if s["combination"] in selected:
                groups.setdefault(s["combination"] or "未分组", []).append(s)
    else:
        groups = {"计算结果": all_students}

    if not any(groups.values()):
        return b""

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, students in groups.items():
        if not students:
            continue
        safe_name = _sanitize_sheet_name(sheet_name)
        ws = wb.create_sheet(title=safe_name)
        # 写表头
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # 写数据
        rows = _build_student_rows(students, use_conversion, columns)
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
        _apply_sheet_styling(ws, columns, rows, use_conversion)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_conversion_report(
    db: Session,
    exam_id: int,
    template_id: int,
) -> bytes:
    """导出赋分报告为 Excel，每个赋分科目一张工作表"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    result = calculate_grades(db, exam_id, use_conversion=True, template_id=template_id)
    report = result.get("conversion_report", {})
    if not report:
        return b""

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center")
    tier_fills = {
        "A": PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"),
        "B": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "C": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "D": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "E": PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid"),
    }

    columns = ["学科", "等级", "比例(%)", "人数", "原始分上限", "原始分下限", "赋分区间", "赋分区间平均分"]

    for subj in sorted(report.keys()):
        tiers = report[subj]
        safe_name = _sanitize_sheet_name(f"{subj}赋分报告")
        ws = wb.create_sheet(title=safe_name)

        # 表头
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # 数据行
        for row_idx, t in enumerate(tiers, 2):
            values = [
                subj,
                t["tier"],
                t["ratio"],
                t["count"],
                t["raw_max"] if t["count"] > 0 else "",
                t["raw_min"] if t["count"] > 0 else "",
                t["score_range"],
                t["converted_avg"] if t["count"] > 0 else "",
            ]
            tier_fill = tier_fills.get(t["tier"], PatternFill())
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = tier_fill
                cell.alignment = center

        # 列宽
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for row_idx in range(2, len(tiers) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 20)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
